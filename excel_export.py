"""
Excel Export Utility for Ceramic QC System
Exports data to Excel templates based on R2-LABO forms
"""

import xlrd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import shutil
import os
from datetime import datetime

class ExcelExporter:
    def __init__(self):
        self.template_dir = "templates"
        self.exports_dir = "exports"
        
        # Create exports directory if it doesn't exist
        if not os.path.exists(self.exports_dir):
            os.makedirs(self.exports_dir)
    
    def export_humidity_data(self, clay_control_data, export_type="combined"):
        """Export humidity control data to Excel template"""
        template_file = os.path.join(self.template_dir, "humidity_template.xls")
        
        if not os.path.exists(template_file):
            raise FileNotFoundError("Humidity template not found")
        
        # Create unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"R2-F1-LABO_Humidite_{timestamp}.xlsx"
        output_path = os.path.join(self.exports_dir, output_filename)
        
        # Convert .xls to .xlsx for easier editing
        temp_xlsx = self._convert_xls_to_xlsx(template_file)
        
        try:
            # Load the workbook
            wb = load_workbook(temp_xlsx)
            ws = wb.active
            
            # Fill in the data based on template structure
            self._fill_humidity_template(ws, clay_control_data)
            
            # Save the file
            wb.save(output_path)
            
            return output_path, output_filename
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
    
    def export_analysis_data(self, clay_control_data, export_type="combined"):
        """Export analysis control data to Excel template"""
        template_file = os.path.join(self.template_dir, "analysis_template.xls")
        
        if not os.path.exists(template_file):
            raise FileNotFoundError("Analysis template not found")
        
        # Create unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"R2-F2-LABO_Analyse_{timestamp}.xlsx"
        output_path = os.path.join(self.exports_dir, output_filename)
        
        # Convert .xls to .xlsx for easier editing
        temp_xlsx = self._convert_xls_to_xlsx(template_file)
        
        try:
            # Load the workbook
            wb = load_workbook(temp_xlsx)
            ws = wb.active
            
            # Fill in the data based on template structure
            self._fill_analysis_template(ws, clay_control_data)
            
            # Save the file
            wb.save(output_path)
            
            return output_path, output_filename
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
    
    def _convert_xls_to_xlsx(self, xls_file):
        """Convert .xls file to .xlsx format"""
        # Read .xls file
        xls_book = xlrd.open_workbook(xls_file)
        
        # Create new .xlsx workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
        
        # Copy each sheet
        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            wb_sheet = wb.create_sheet(title=sheet_name)
            
            # Copy data
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row, col)
                    wb_sheet.cell(row=row+1, column=col+1, value=cell_value)
        
        # Set first sheet as active
        if wb.sheetnames:
            wb.active = wb[wb.sheetnames[0]]
        
        # Save temporary file
        temp_file = xls_file.replace('.xls', '_temp.xlsx')
        wb.save(temp_file)
        
        return temp_file
    
    def _fill_humidity_template(self, worksheet, data):
        """Fill humidity template with actual data"""
        # Find data entry areas and fill them
        date_str = data.get('date', datetime.now()).strftime("%d/%m/%Y")
        shift = data.get('shift', '')
        controller = data.get('controller', 'Non spécifié')
        
        # Set controller name (around row 7)
        try:
            worksheet.cell(row=7, column=1, value=f"Contrôleur : {controller}")
            worksheet.cell(row=9, column=8, value=f"Contrôleur : {controller}")
        except:
            pass
        
        # Fill humidity data starting around row 10-12
        row_start = 11
        
        # Humidity before prep
        if data.get('humidity_before_prep'):
            worksheet.cell(row=row_start, column=2, value=date_str)
            worksheet.cell(row=row_start, column=3, value=shift)
            worksheet.cell(row=row_start, column=4, value=data['humidity_before_prep'])
            worksheet.cell(row=row_start, column=5, value=data.get('measurement_time_1', ''))
            worksheet.cell(row=row_start, column=6, value="2.5% ≤ H ≤ 4.1%")
            status = "CONFORME" if 2.5 <= data['humidity_before_prep'] <= 4.1 else "NON CONFORME"
            worksheet.cell(row=row_start, column=7, value=status)
        
        # Humidity after sieving
        if data.get('humidity_after_sieving'):
            worksheet.cell(row=row_start+1, column=2, value=date_str)
            worksheet.cell(row=row_start+1, column=3, value=shift)
            worksheet.cell(row=row_start+1, column=4, value=data['humidity_after_sieving'])
            worksheet.cell(row=row_start+1, column=5, value=data.get('measurement_time_2', ''))
            worksheet.cell(row=row_start+1, column=6, value="2% ≤ H ≤ 3.5%")
            status = "CONFORME" if 2.0 <= data['humidity_after_sieving'] <= 3.5 else "NON CONFORME"
            worksheet.cell(row=row_start+1, column=7, value=status)
        
        # Humidity after prep
        if data.get('humidity_after_prep'):
            worksheet.cell(row=row_start+2, column=2, value=date_str)
            worksheet.cell(row=row_start+2, column=3, value=shift)
            worksheet.cell(row=row_start+2, column=4, value=data['humidity_after_prep'])
            worksheet.cell(row=row_start+2, column=5, value=data.get('measurement_time_3', ''))
            worksheet.cell(row=row_start+2, column=6, value="5.3% ≤ H ≤ 6.3%")
            status = "CONFORME" if 5.3 <= data['humidity_after_prep'] <= 6.3 else "NON CONFORME"
            worksheet.cell(row=row_start+2, column=7, value=status)
        
        # Add notes if any
        if data.get('notes'):
            worksheet.cell(row=row_start+5, column=1, value=f"Notes: {data['notes']}")
    
    def _fill_analysis_template(self, worksheet, data):
        """Fill analysis template with actual data"""
        # Find data entry areas and fill them
        date_str = data.get('date', datetime.now()).strftime("%d/%m/%Y")
        shift = data.get('shift', '')
        controller = data.get('controller', 'Non spécifié')
        
        # Fill header info
        try:
            worksheet.cell(row=3, column=1, value=f"Contrôleur : {controller}")
            worksheet.cell(row=4, column=1, value=f"Date : {date_str}")
            worksheet.cell(row=5, column=1, value=f"Équipe : {shift}")
        except:
            pass
        
        # Fill granulometry data starting around row 10
        row_start = 11
        
        if data.get('granulometry_refusal'):
            worksheet.cell(row=row_start, column=1, value="Granulométrie CaCO₃")
            worksheet.cell(row=row_start, column=2, value=f"{data['granulometry_refusal']}%")
            worksheet.cell(row=row_start, column=3, value="10% ≤ Refus ≤ 20%")
            status = "CONFORME" if 10 <= data['granulometry_refusal'] <= 20 else "NON CONFORME"
            worksheet.cell(row=row_start, column=4, value=status)
            worksheet.cell(row=row_start, column=5, value=date_str)
        
        # Fill calcium carbonate data
        if data.get('calcium_carbonate'):
            worksheet.cell(row=row_start+1, column=1, value="Carbonate CaCO₃")
            worksheet.cell(row=row_start+1, column=2, value=f"{data['calcium_carbonate']}%")
            worksheet.cell(row=row_start+1, column=3, value="15% ≤ CaCO₃ ≤ 25%")
            status = "CONFORME" if 15 <= data['calcium_carbonate'] <= 25 else "NON CONFORME"
            worksheet.cell(row=row_start+1, column=4, value=status)
            worksheet.cell(row=row_start+1, column=5, value=date_str)
        
        # Add notes if any
        if data.get('notes'):
            worksheet.cell(row=row_start+4, column=1, value=f"Notes: {data['notes']}")

    def get_exports_list(self):
        """Get list of exported files"""
        if not os.path.exists(self.exports_dir):
            return []
        
        files = []
        for filename in os.listdir(self.exports_dir):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(self.exports_dir, filename)
                stat = os.stat(filepath)
                files.append({
                    'filename': filename,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime)
                })
        
        return sorted(files, key=lambda x: x['modified'], reverse=True)