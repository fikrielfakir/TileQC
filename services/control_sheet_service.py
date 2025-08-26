from models import db, ControlSheet, ScheduledControl, OptimizedMeasurement, ControlParameter, ControlStage
from services.scheduling_service import SchedulingService
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io

class ControlSheetService:
    """Service for generating automated control sheets"""
    
    @staticmethod
    def generate_daily_control_sheet(target_date=None, shift=None, format_type='excel'):
        """Generate daily control sheet with scheduled controls and actual measurements"""
        
        if not target_date:
            target_date = date.today()
        
        # Get scheduled controls for the date
        scheduled_controls = SchedulingService.get_daily_schedule(target_date, shift)
        
        # Get actual measurements for the date
        measurements_query = OptimizedMeasurement.query.filter_by(measurement_date=target_date)
        if shift:
            measurements_query = measurements_query.filter_by(shift=shift)
        measurements = measurements_query.all()
        
        # Organize data by parameter
        control_data = {}
        for control in scheduled_controls:
            param_id = control.parameter_id
            if param_id not in control_data:
                control_data[param_id] = {
                    'parameter': control.parameter,
                    'scheduled': [],
                    'measurements': []
                }
            control_data[param_id]['scheduled'].append(control)
        
        for measurement in measurements:
            param_id = measurement.parameter_id
            if param_id not in control_data:
                control_data[param_id] = {
                    'parameter': measurement.parameter,
                    'scheduled': [],
                    'measurements': []
                }
            control_data[param_id]['measurements'].append(measurement)
        
        if format_type == 'excel':
            return ControlSheetService._generate_excel_control_sheet(target_date, control_data, shift)
        else:
            return ControlSheetService._generate_pdf_control_sheet(target_date, control_data, shift)
    
    @staticmethod
    def _generate_excel_control_sheet(target_date, control_data, shift=None):
        """Generate Excel control sheet"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title
        title = f"FICHE DE CONTRÔLE - {target_date.strftime('%d/%m/%Y')}"
        if shift:
            title += f" - Équipe {shift}"
        
        ws.title = title[:31]  # Excel sheet name limit
        
        # Header styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        
        # Title
        ws['A1'] = title
        ws['A1'].font = header_font
        ws.merge_cells('A1:J1')
        
        # Headers row
        row = 3
        headers = [
            'Étape', 'Paramètre', 'Spécification', 'Fréquence',
            'Heures Prévues', 'Mesures Réalisées', 'Conformité',
            'Écarts', 'NC', 'Observations'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        row += 1
        
        # Data rows
        for param_id, data in control_data.items():
            parameter = data['parameter']
            scheduled = data['scheduled']
            measurements = data['measurements']
            
            # Parameter info
            ws.cell(row=row, column=1, value=parameter.stage.name)
            ws.cell(row=row, column=2, value=parameter.name)
            ws.cell(row=row, column=3, value=parameter.specification)
            ws.cell(row=row, column=4, value=f"{parameter.frequency_per_day}x/jour")
            
            # Scheduled times
            scheduled_times = ', '.join([s.scheduled_time.strftime('%H:%M') for s in scheduled])
            ws.cell(row=row, column=5, value=scheduled_times)
            
            # Actual measurements
            if measurements:
                measurement_times = ', '.join([m.measurement_time.strftime('%H:%M') for m in measurements])
                ws.cell(row=row, column=6, value=measurement_times)
                
                # Conformity status
                conforming_count = sum(1 for m in measurements if m.is_conforming)
                total_count = len(measurements)
                conformity_text = f"{conforming_count}/{total_count}"
                
                cell = ws.cell(row=row, column=7, value=conformity_text)
                if conforming_count < total_count:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                # Deviations
                deviations = [str(m.deviation_percentage) + '%' for m in measurements 
                             if m.deviation_percentage is not None]
                ws.cell(row=row, column=8, value=', '.join(deviations))
                
                # NC numbers
                nc_numbers = [m.nc_number for m in measurements if m.nc_number]
                ws.cell(row=row, column=9, value=', '.join(nc_numbers))
                
                # Observations
                observations = [m.observations for m in measurements if m.observations]
                ws.cell(row=row, column=10, value='; '.join(observations))
            else:
                ws.cell(row=row, column=6, value="Aucune mesure")
                cell = ws.cell(row=row, column=7, value="Manquant")
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row += 1
        
        # Summary section
        row += 2
        ws.cell(row=row, column=1, value="RÉSUMÉ").font = subheader_font
        row += 1
        
        total_params = len(control_data)
        completed_params = sum(1 for data in control_data.values() if data['measurements'])
        conforming_params = sum(1 for data in control_data.values() 
                               if data['measurements'] and all(m.is_conforming for m in data['measurements']))
        
        ws.cell(row=row, column=1, value=f"Paramètres contrôlés: {completed_params}/{total_params}")
        row += 1
        ws.cell(row=row, column=1, value=f"Paramètres conformes: {conforming_params}/{completed_params}")
        row += 1
        ws.cell(row=row, column=1, value=f"Taux de conformité: {(conforming_params/completed_params*100):.1f}%" if completed_params > 0 else "N/A")
        
        # Auto-adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        filename = f"Fiche_Controle_{target_date.strftime('%Y%m%d')}"
        if shift:
            filename += f"_{shift.replace('-', '')}"
        filename += ".xlsx"
        
        return {
            'success': True,
            'buffer': buffer,
            'filename': filename,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    
    @staticmethod
    def generate_shift_control_sheet(target_date, shift):
        """Generate control sheet for specific shift"""
        return ControlSheetService.generate_daily_control_sheet(target_date, shift)
    
    @staticmethod
    def generate_weekly_control_sheet(start_date):
        """Generate weekly control sheet summary"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Semaine_{start_date.strftime('%Y%m%d')}"
        
        # Header
        ws['A1'] = f"RAPPORT HEBDOMADAIRE - Semaine du {start_date.strftime('%d/%m/%Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Daily summary for the week
        row = 3
        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            day_name = days[i]
            
            # Get daily summary
            summary = SchedulingService.get_schedule_summary(current_date)
            
            ws.cell(row=row, column=1, value=f"{day_name} {current_date.strftime('%d/%m')}")
            ws.cell(row=row, column=2, value=summary['total'])
            ws.cell(row=row, column=3, value=summary['by_status']['completed'])
            ws.cell(row=row, column=4, value=summary['by_status']['pending'])
            ws.cell(row=row, column=5, value=summary['by_status']['overdue'])
            
            completion_rate = (summary['by_status']['completed'] / summary['total'] * 100) if summary['total'] > 0 else 0
            ws.cell(row=row, column=6, value=f"{completion_rate:.1f}%")
            
            row += 1
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Rapport_Hebdo_{start_date.strftime('%Y%m%d')}.xlsx"
        
        return {
            'success': True,
            'buffer': buffer,
            'filename': filename,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    
    @staticmethod
    def save_control_sheet(sheet_type, target_date, generated_by, file_path, shift=None, stage_id=None):
        """Save control sheet record to database"""
        
        control_sheet = ControlSheet(
            sheet_type=sheet_type,
            reference_date=target_date,
            shift=shift,
            stage_id=stage_id,
            generated_by=generated_by,
            file_path=file_path,
            status='final'
        )
        
        db.session.add(control_sheet)
        db.session.commit()
        
        return control_sheet.id