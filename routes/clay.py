from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from forms import ClayControlForm, HumidityBeforePrepForm, HumidityAfterSievingForm, HumidityAfterPrepForm, GranulometryForm, CalciumCarbonateForm, CombinedHumidityForm, CombinedAnalysisForm
from models import ClayControl, DryerControl, PressControl
from app import db
from datetime import date, datetime
from excel_export import ExcelExporter
import os
import io

clay_bp = Blueprint('clay', __name__)

@clay_bp.route('/')
@login_required
def clay_controls():
    page = request.args.get('page', 1, type=int)
    controls = ClayControl.query.order_by(ClayControl.date.desc()).paginate(
        page=page, per_page=20, error_out=False)
    return render_template('clay/clay_control.html', controls=controls)

@clay_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_clay_control():
    form = ClayControlForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time_1.data
        clay_control.measurement_time_2 = form.measurement_time_2.data
        clay_control.humidity_before_prep = form.humidity_before_prep.data
        clay_control.humidity_after_sieving = form.humidity_after_sieving.data
        clay_control.humidity_after_prep = form.humidity_after_prep.data
        clay_control.granulometry_refusal = form.granulometry_refusal.data
        clay_control.calcium_carbonate = form.calcium_carbonate.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        clay_control.compliance_status = 'compliant'  # Default
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('Enregistrement du contrôle argile ajouté avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/clay_control.html', form=form)

@clay_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_clay_control(id):
    clay_control = ClayControl.query.get_or_404(id)
    form = ClayControlForm(obj=clay_control)
    
    if form.validate_on_submit():
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time_1.data
        clay_control.measurement_time_2 = form.measurement_time_2.data
        clay_control.humidity_before_prep = form.humidity_before_prep.data
        clay_control.humidity_after_sieving = form.humidity_after_sieving.data
        clay_control.humidity_after_prep = form.humidity_after_prep.data
        clay_control.granulometry_refusal = form.granulometry_refusal.data
        clay_control.calcium_carbonate = form.calcium_carbonate.data
        clay_control.notes = form.notes.data
        
        db.session.commit()
        
        flash('Enregistrement du contrôle argile mis à jour avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/clay_control.html', form=form, edit=True, control=clay_control)

@clay_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_clay_control(id):
    if current_user.role not in ['admin', 'quality_manager']:
        flash('Permissions insuffisantes', 'error')
        return redirect(url_for('clay.clay_controls'))
    
    clay_control = ClayControl.query.get_or_404(id)
    db.session.delete(clay_control)
    db.session.commit()
    
    flash('Enregistrement du contrôle argile supprimé avec succès', 'success')
    return redirect(url_for('clay.clay_controls'))

@clay_bp.route('/api/trend/<parameter>')
@login_required
def clay_trend_api(parameter):
    from utils.helpers import get_control_chart_data
    data = get_control_chart_data(ClayControl, parameter, days=30)
    return jsonify(data)

# Separate routes for each clay sub-control

@clay_bp.route('/humidity-before-prep', methods=['GET', 'POST'])
@login_required
def humidity_before_prep():
    form = HumidityBeforePrepForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time.data
        clay_control.humidity_before_prep = form.humidity_before_prep.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance
        humidity_value = form.humidity_before_prep.data
        if humidity_value is not None and (humidity_value < 2.5 or humidity_value > 4.1):
            clay_control.compliance_status = 'non_compliant'
            flash(f'⚠️ Valeur hors spécification ({humidity_value}%) - F.N.C. requis (Spéc: 2.5%-4.1%)', 'warning')
        else:
            clay_control.compliance_status = 'compliant'
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('Humidité trémie générale enregistrée avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/humidity_before_prep.html', form=form)

@clay_bp.route('/humidity-after-sieving', methods=['GET', 'POST'])
@login_required
def humidity_after_sieving():
    form = HumidityAfterSievingForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time.data
        clay_control.humidity_after_sieving = form.humidity_after_sieving.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance
        humidity_value = form.humidity_after_sieving.data
        if humidity_value is not None and (humidity_value < 2.0 or humidity_value > 3.5):
            clay_control.compliance_status = 'non_compliant'
            flash(f'⚠️ Valeur hors spécification ({humidity_value}%) - F.N.C. requis (Spéc: 2%-3.5%)', 'warning')
        else:
            clay_control.compliance_status = 'compliant'
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('Humidité après tamisage enregistrée avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/humidity_after_sieving.html', form=form)

@clay_bp.route('/humidity-after-prep', methods=['GET', 'POST'])
@login_required
def humidity_after_prep():
    form = HumidityAfterPrepForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time.data
        clay_control.humidity_after_prep = form.humidity_after_prep.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance
        humidity_value = form.humidity_after_prep.data
        if humidity_value is not None and (humidity_value < 5.3 or humidity_value > 6.3):
            clay_control.compliance_status = 'non_compliant'
            flash(f'⚠️ Valeur hors spécification ({humidity_value}%) - F.N.C. requis (Spéc: 5.3%-6.3%)', 'warning')
        else:
            clay_control.compliance_status = 'compliant'
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('Humidité niveau silo enregistrée avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/humidity_after_prep.html', form=form)

@clay_bp.route('/granulometry', methods=['GET', 'POST'])
@login_required
def granulometry():
    form = GranulometryForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time.data
        clay_control.granulometry_refusal = form.granulometry_refusal.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance
        granulo_value = form.granulometry_refusal.data
        if granulo_value is not None and (granulo_value < 10 or granulo_value > 20):
            clay_control.compliance_status = 'non_compliant'
            flash(f'⚠️ Valeur hors spécification ({granulo_value}%) - F.N.C. requis (Spéc: 10%-20%)', 'warning')
        else:
            clay_control.compliance_status = 'compliant'
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('Granulométrie enregistrée avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/granulometry.html', form=form)

@clay_bp.route('/calcium-carbonate', methods=['GET', 'POST'])
@login_required
def calcium_carbonate():
    form = CalciumCarbonateForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time.data
        clay_control.calcium_carbonate = form.calcium_carbonate.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance
        calcium_value = form.calcium_carbonate.data
        if calcium_value is not None and (calcium_value < 15 or calcium_value > 25):
            clay_control.compliance_status = 'non_compliant'
            flash(f'⚠️ Valeur hors spécification ({calcium_value}%) - F.N.C. requis (Spéc: 15%-25%)', 'warning')
        else:
            clay_control.compliance_status = 'compliant'
        
        db.session.add(clay_control)
        db.session.commit()
        
        flash('% Chaux CaCO₃ enregistré avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/calcium_carbonate.html', form=form)

# Combined forms routes as requested by user
@clay_bp.route('/combined-humidity', methods=['GET', 'POST'])
@login_required
def combined_humidity():
    form = CombinedHumidityForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.measurement_time_1 = form.measurement_time_1.data
        clay_control.measurement_time_2 = form.measurement_time_2.data
        clay_control.humidity_before_prep = form.humidity_before_prep.data
        clay_control.humidity_after_sieving = form.humidity_after_sieving.data
        clay_control.humidity_after_prep = form.humidity_after_prep.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance for all humidity values
        compliance_issues = []
        overall_compliance = 'compliant'
        
        if form.humidity_before_prep.data:
            if form.humidity_before_prep.data < 2.5 or form.humidity_before_prep.data > 4.1:
                compliance_issues.append(f'Trémie générale ({form.humidity_before_prep.data}%) - Spéc: 2.5%-4.1%')
                overall_compliance = 'non_compliant'
                
        if form.humidity_after_sieving.data:
            if form.humidity_after_sieving.data < 2.0 or form.humidity_after_sieving.data > 3.5:
                compliance_issues.append(f'Après tamisage ({form.humidity_after_sieving.data}%) - Spéc: 2%-3.5%')
                overall_compliance = 'non_compliant'
                
        if form.humidity_after_prep.data:
            if form.humidity_after_prep.data < 5.3 or form.humidity_after_prep.data > 6.3:
                compliance_issues.append(f'Niveau silo ({form.humidity_after_prep.data}%) - Spéc: 5.3%-6.3%')
                overall_compliance = 'non_compliant'
        
        clay_control.compliance_status = overall_compliance
        
        db.session.add(clay_control)
        db.session.commit()
        
        if compliance_issues:
            flash(f'⚠️ Valeurs hors spécification: {", ".join(compliance_issues)} - F.N.C. requis', 'warning')
        
        # Check if export to Excel was requested
        if request.form.get('export_excel'):
            try:
                exporter = ExcelExporter()
                export_data = {
                    'date': form.date.data,
                    'shift': form.shift.data,
                    'controller': current_user.username,
                    'humidity_before_prep': form.humidity_before_prep.data,
                    'humidity_after_sieving': form.humidity_after_sieving.data,
                    'humidity_after_prep': form.humidity_after_prep.data,
                    'measurement_time_1': form.measurement_time_1.data,
                    'measurement_time_2': form.measurement_time_2.data,
                    'measurement_time_3': form.measurement_time_3.data,
                    'notes': form.notes.data
                }
                file_path, filename = exporter.export_humidity_data(export_data)
                flash(f'✅ Données exportées vers Excel: {filename}', 'success')
                return send_file(file_path, as_attachment=True, download_name=filename)
            except Exception as e:
                flash(f'❌ Erreur lors de l\'export Excel: {str(e)}', 'error')
        
        flash('Contrôles d\'humidité enregistrés avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/combined_humidity.html', form=form)

@clay_bp.route('/combined-analysis', methods=['GET', 'POST'])
@login_required
def combined_analysis():
    form = CombinedAnalysisForm()
    
    if form.validate_on_submit():
        clay_control = ClayControl()
        clay_control.date = form.date.data
        clay_control.shift = form.shift.data
        clay_control.granulometry_refusal = form.granulometry_refusal.data
        clay_control.calcium_carbonate = form.calcium_carbonate.data
        clay_control.notes = form.notes.data
        clay_control.controller_id = current_user.id
        
        # Check compliance for both analysis values
        compliance_issues = []
        overall_compliance = 'compliant'
        
        if form.granulometry_refusal.data:
            if form.granulometry_refusal.data < 10 or form.granulometry_refusal.data > 20:
                compliance_issues.append(f'Granulométrie ({form.granulometry_refusal.data}%) - Spéc: 10%-20%')
                overall_compliance = 'non_compliant'
                
        if form.calcium_carbonate.data:
            if form.calcium_carbonate.data < 15 or form.calcium_carbonate.data > 25:
                compliance_issues.append(f'CaCO₃ ({form.calcium_carbonate.data}%) - Spéc: 15%-25%')
                overall_compliance = 'non_compliant'
        
        clay_control.compliance_status = overall_compliance
        
        db.session.add(clay_control)
        db.session.commit()
        
        if compliance_issues:
            flash(f'⚠️ Valeurs hors spécification: {", ".join(compliance_issues)} - F.N.C. requis', 'warning')
        
        # Check if export to Excel was requested
        if request.form.get('export_excel'):
            try:
                exporter = ExcelExporter()
                export_data = {
                    'date': form.date.data,
                    'shift': form.shift.data,
                    'controller': current_user.username,
                    'granulometry_refusal': form.granulometry_refusal.data,
                    'calcium_carbonate': form.calcium_carbonate.data,
                    'granulometry_time': form.granulometry_time.data,
                    'calcium_time': form.calcium_time.data,
                    'notes': form.notes.data
                }
                file_path, filename = exporter.export_analysis_data(export_data)
                flash(f'✅ Données exportées vers Excel: {filename}', 'success')
                return send_file(file_path, as_attachment=True, download_name=filename)
            except Exception as e:
                flash(f'❌ Erreur lors de l\'export Excel: {str(e)}', 'error')
        
        flash('Analyses enregistrées avec succès', 'success')
        return redirect(url_for('clay.clay_controls'))
    
    return render_template('clay/combined_analysis.html', form=form)

# Control Sheet Generation Routes
@clay_bp.route('/generate-humidity-sheet')
@login_required
def generate_humidity_sheet():
    """Generate Fiche Contrôle Humidité"""
    return render_template('clay/humidity_control_sheet.html')

@clay_bp.route('/generate-clay-sheet') 
@login_required
def generate_clay_sheet():
    """Generate Fiche de Contrôle Argile"""
    return render_template('clay/clay_control_sheet.html')

@clay_bp.route('/control_sheet', methods=['GET', 'POST'])
@login_required
def control_sheet():
    """Generate PDF/Excel with real data for selected day"""
    
    if request.method == 'POST':
        selected_date = request.form.get('selected_date')
        export_format = request.form.get('export_format', 'pdf')
        
        if not selected_date:
            flash('Veuillez sélectionner une date', 'error')
            return redirect(url_for('clay.control_sheet'))
        
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            flash('Format de date invalide', 'error')
            return redirect(url_for('clay.control_sheet'))
        
        # Retrieve humidity data for the selected date
        clay_controls = ClayControl.query.filter_by(date=selected_date).all()
        dryer_controls = DryerControl.query.filter_by(date=selected_date).all()
        
        # Organize the data by humidity type
        humidity_data = {
            'date': selected_date,
            'tremie_generale': [],
            'apres_tamisage': [],
            'niveau_silo': [],
            'argile_presse': [],
            'residuelle_sechoir': []
        }
        
        # Extract clay control humidity data
        for control in clay_controls:
            if control.humidity_before_prep is not None:
                humidity_data['tremie_generale'].append({
                    'time': control.measurement_time_1,
                    'value': control.humidity_before_prep,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 2.5,
                    'spec_max': 4.1,
                    'compliant': 2.5 <= control.humidity_before_prep <= 4.1 if control.humidity_before_prep else False
                })
            
            if control.humidity_after_sieving is not None:
                humidity_data['apres_tamisage'].append({
                    'time': control.measurement_time_1,
                    'value': control.humidity_after_sieving,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 2.0,
                    'spec_max': 3.5,
                    'compliant': 2.0 <= control.humidity_after_sieving <= 3.5 if control.humidity_after_sieving else False
                })
            
            if control.humidity_after_prep is not None:
                humidity_data['niveau_silo'].append({
                    'time': control.measurement_time_2,
                    'value': control.humidity_after_prep,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 5.3,
                    'spec_max': 6.3,
                    'compliant': 5.3 <= control.humidity_after_prep <= 6.3 if control.humidity_after_prep else False
                })
        
        # Extract dryer control data for residual humidity
        for control in dryer_controls:
            if hasattr(control, 'residual_humidity') and control.residual_humidity is not None:
                humidity_data['residuelle_sechoir'].append({
                    'time': getattr(control, 'measurement_time', None),
                    'value': control.residual_humidity,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 0.1,
                    'spec_max': 1.5,
                    'compliant': 0.1 <= control.residual_humidity <= 1.5 if control.residual_humidity else False
                })
        
        if export_format == 'pdf':
            return generate_humidity_pdf(humidity_data)
        else:
            return generate_humidity_excel(humidity_data)
    
    return render_template('clay/control_sheet_selector.html')

def generate_humidity_pdf(data):
    """Generate PDF with humidity control data"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
    except ImportError:
        # Fallback to simple text-based PDF
        flash('ReportLab non disponible. Génération PDF simplifiée.', 'warning')
        return generate_simple_pdf(data)
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, f"FICHE CONTRÔLE HUMIDITÉ - {data['date'].strftime('%d/%m/%Y')}")
    
    y_position = height - 100
    
    # Function to draw section
    def draw_humidity_section(title, section_data, spec_range):
        nonlocal y_position
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y_position, title)
        y_position -= 30
        
        if section_data:
            for item in section_data:
                p.setFont("Helvetica", 10)
                time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
                status = "✓" if item['compliant'] else "✗"
                text = f"  {time_str} - {item['value']}% - {item['controller']} - Spéc: {spec_range} - {status}"
                p.drawString(70, y_position, text)
                y_position -= 15
        else:
            p.setFont("Helvetica-Oblique", 10)
            p.drawString(70, y_position, "Aucune mesure disponible")
            y_position -= 15
        
        y_position -= 10
    
    # Draw sections
    draw_humidity_section("1. HUMIDITÉ TRÉMIE GÉNÉRALE", data['tremie_generale'], "2,5% - 4,1%")
    draw_humidity_section("2. HUMIDITÉ APRÈS TAMISAGE", data['apres_tamisage'], "2% - 3,5%")
    draw_humidity_section("3. HUMIDITÉ NIVEAU SILO", data['niveau_silo'], "5,3% - 6,3%")
    draw_humidity_section("4. HUMIDITÉ RÉSIDUELLE SÉCHOIR", data['residuelle_sechoir'], "0,1% - 1,5%")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

def generate_humidity_excel(data):
    """Generate Excel with humidity control data"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        flash('OpenPyXL non disponible. Génération Excel non supportée.', 'error')
        return redirect(url_for('clay.control_sheet'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Humidité_{data['date'].strftime('%Y%m%d')}"
    
    # Headers
    ws['A1'] = 'FICHE CONTRÔLE HUMIDITÉ'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Date: {data['date'].strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True)
    
    row = 4
    
    # Function to add section
    def add_section(title, section_data, spec_range):
        nonlocal row
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ['Heure', 'Valeur (%)', 'Contrôleur', 'Spécifications', 'Conforme']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        row += 1
        
        if section_data:
            for item in section_data:
                time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
                ws[f'A{row}'] = time_str
                ws[f'B{row}'] = item['value']
                ws[f'C{row}'] = item['controller']
                ws[f'D{row}'] = spec_range
                ws[f'E{row}'] = 'OUI' if item['compliant'] else 'NON'
                
                # Color coding
                if not item['compliant']:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                row += 1
        else:
            ws[f'A{row}'] = 'Aucune mesure disponible'
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
        
        row += 1
    
    # Add sections
    add_section("1. HUMIDITÉ TRÉMIE GÉNÉRALE", data['tremie_generale'], "2,5% - 4,1%")
    add_section("2. HUMIDITÉ APRÈS TAMISAGE", data['apres_tamisage'], "2% - 3,5%")
    add_section("3. HUMIDITÉ NIVEAU SILO", data['niveau_silo'], "5,3% - 6,3%")
    add_section("4. HUMIDITÉ RÉSIDUELLE SÉCHOIR", data['residuelle_sechoir'], "0,1% - 1,5%")
    
    # Auto-adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generate_simple_pdf(data):
    """Generate simple text-based PDF fallback"""
    buffer = io.StringIO()
    content = f"""FICHE CONTRÔLE HUMIDITÉ
Date: {data['date'].strftime('%d/%m/%Y')}

1. HUMIDITÉ TRÉMIE GÉNÉRALE (Spéc: 2,5% - 4,1%)
"""
    
    if data['tremie_generale']:
        for item in data['tremie_generale']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n2. HUMIDITÉ APRÈS TAMISAGE (Spéc: 2% - 3,5%)\n"
    if data['apres_tamisage']:
        for item in data['apres_tamisage']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n3. HUMIDITÉ NIVEAU SILO (Spéc: 5,3% - 6,3%)\n"
    if data['niveau_silo']:
        for item in data['niveau_silo']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n4. HUMIDITÉ RÉSIDUELLE SÉCHOIR (Spéc: 0,1% - 1,5%)\n"
    if data['residuelle_sechoir']:
        for item in data['residuelle_sechoir']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"
    
    # Create a simple text file as fallback
    buffer = io.BytesIO(content.encode('utf-8'))
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.txt"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='text/plain'
    )

