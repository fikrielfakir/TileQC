from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import ClayControl, DryerControl, PressControl
from app import db
from datetime import date, datetime
import io

control_sheet_bp = Blueprint('control_sheet', __name__)

@control_sheet_bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    """Generate PDF/Excel with real data for selected day"""
    
    if request.method == 'POST':
        selected_date = request.form.get('selected_date')
        export_format = request.form.get('export_format', 'pdf')
        
        if not selected_date:
            flash('Veuillez sélectionner une date', 'error')
            return redirect(url_for('control_sheet.index'))
        
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            flash('Format de date invalide', 'error')
            return redirect(url_for('control_sheet.index'))
        
        # Retrieve humidity data for the selected date
        clay_controls = ClayControl.query.filter_by(date=selected_date).all()
        dryer_controls = DryerControl.query.filter_by(date=selected_date).all()
        
        # Organize the data by humidity type
        humidity_data = {
            'date': selected_date,
            'tremie_generale': [],
            'apres_tamisage': [],
            'niveau_silo': [],
            'argile_presse': [],
            'residuelle_sechoir': []
        }
        
        # Extract clay control humidity data
        for control in clay_controls:
            if control.humidity_before_prep is not None:
                humidity_data['tremie_generale'].append({
                    'time': control.measurement_time_1,
                    'value': control.humidity_before_prep,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 2.5,
                    'spec_max': 4.1,
                    'compliant': 2.5 <= control.humidity_before_prep <= 4.1 if control.humidity_before_prep else False
                })
            
            if control.humidity_after_sieving is not None:
                humidity_data['apres_tamisage'].append({
                    'time': control.measurement_time_1,
                    'value': control.humidity_after_sieving,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 2.0,
                    'spec_max': 3.5,
                    'compliant': 2.0 <= control.humidity_after_sieving <= 3.5 if control.humidity_after_sieving else False
                })
            
            if control.humidity_after_prep is not None:
                humidity_data['niveau_silo'].append({
                    'time': control.measurement_time_2,
                    'value': control.humidity_after_prep,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 5.3,
                    'spec_max': 6.3,
                    'compliant': 5.3 <= control.humidity_after_prep <= 6.3 if control.humidity_after_prep else False
                })
        
        # Extract dryer control data for residual humidity
        for control in dryer_controls:
            if hasattr(control, 'residual_humidity') and control.residual_humidity is not None:
                humidity_data['residuelle_sechoir'].append({
                    'time': getattr(control, 'measurement_time', None),
                    'value': control.residual_humidity,
                    'controller': control.controller.username if control.controller else '',
                    'spec_min': 0.1,
                    'spec_max': 1.5,
                    'compliant': 0.1 <= control.residual_humidity <= 1.5 if control.residual_humidity else False
                })
        
        if export_format == 'pdf':
            return generate_humidity_pdf(humidity_data)
        else:
            return generate_humidity_excel(humidity_data)
    
    return render_template('control_sheet/index.html')

@control_sheet_bp.route('/generate-humidity-sheet')
@login_required
def generate_humidity_sheet():
    """Generate Fiche Contrôle Humidité"""
    return render_template('control_sheet/humidity_control_sheet.html')

@control_sheet_bp.route('/generate-clay-sheet') 
@login_required
def generate_clay_sheet():
    """Generate Fiche de Contrôle Argile"""
    return render_template('control_sheet/clay_control_sheet.html')

def generate_humidity_pdf(data):
    """Generate PDF with humidity control data"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
    except ImportError:
        # Fallback to simple text-based PDF
        flash('ReportLab non disponible. Génération PDF simplifiée.', 'warning')
        return generate_simple_pdf(data)
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, f"FICHE CONTRÔLE HUMIDITÉ - {data['date'].strftime('%d/%m/%Y')}")
    
    y_position = height - 100
    
    # Function to draw section
    def draw_humidity_section(title, section_data, spec_range):
        nonlocal y_position
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y_position, title)
        y_position -= 30
        
        if section_data:
            for item in section_data:
                p.setFont("Helvetica", 10)
                time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
                status = "✓" if item['compliant'] else "✗"
                text = f"  {time_str} - {item['value']}% - {item['controller']} - Spéc: {spec_range} - {status}"
                p.drawString(70, y_position, text)
                y_position -= 15
        else:
            p.setFont("Helvetica-Oblique", 10)
            p.drawString(70, y_position, "Aucune mesure disponible")
            y_position -= 15
        
        y_position -= 10
    
    # Draw sections
    draw_humidity_section("1. HUMIDITÉ TRÉMIE GÉNÉRALE", data['tremie_generale'], "2,5% - 4,1%")
    draw_humidity_section("2. HUMIDITÉ APRÈS TAMISAGE", data['apres_tamisage'], "2% - 3,5%")
    draw_humidity_section("3. HUMIDITÉ NIVEAU SILO", data['niveau_silo'], "5,3% - 6,3%")
    draw_humidity_section("4. HUMIDITÉ RÉSIDUELLE SÉCHOIR", data['residuelle_sechoir'], "0,1% - 1,5%")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

def generate_humidity_excel(data):
    """Generate Excel with humidity control data"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        flash('OpenPyXL non disponible. Génération Excel non supportée.', 'error')
        return redirect(url_for('control_sheet.index'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Humidité_{data['date'].strftime('%Y%m%d')}"
    
    # Headers
    ws['A1'] = 'FICHE CONTRÔLE HUMIDITÉ'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Date: {data['date'].strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True)
    
    row = 4
    
    # Function to add section
    def add_section(title, section_data, spec_range):
        nonlocal row
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ['Heure', 'Valeur (%)', 'Contrôleur', 'Spécifications', 'Conforme']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        row += 1
        
        if section_data:
            for item in section_data:
                time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
                ws[f'A{row}'] = time_str
                ws[f'B{row}'] = item['value']
                ws[f'C{row}'] = item['controller']
                ws[f'D{row}'] = spec_range
                ws[f'E{row}'] = 'OUI' if item['compliant'] else 'NON'
                
                # Color coding
                if not item['compliant']:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                row += 1
        else:
            ws[f'A{row}'] = 'Aucune mesure disponible'
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
        
        row += 1
    
    # Add sections
    add_section("1. HUMIDITÉ TRÉMIE GÉNÉRALE", data['tremie_generale'], "2,5% - 4,1%")
    add_section("2. HUMIDITÉ APRÈS TAMISAGE", data['apres_tamisage'], "2% - 3,5%")
    add_section("3. HUMIDITÉ NIVEAU SILO", data['niveau_silo'], "5,3% - 6,3%")
    add_section("4. HUMIDITÉ RÉSIDUELLE SÉCHOIR", data['residuelle_sechoir'], "0,1% - 1,5%")
    
    # Auto-adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generate_simple_pdf(data):
    """Generate simple text-based PDF fallback"""
    buffer = io.StringIO()
    content = f"""FICHE CONTRÔLE HUMIDITÉ
Date: {data['date'].strftime('%d/%m/%Y')}

1. HUMIDITÉ TRÉMIE GÉNÉRALE (Spéc: 2,5% - 4,1%)
"""
    
    if data['tremie_generale']:
        for item in data['tremie_generale']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n2. HUMIDITÉ APRÈS TAMISAGE (Spéc: 2% - 3,5%)\n"
    if data['apres_tamisage']:
        for item in data['apres_tamisage']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n3. HUMIDITÉ NIVEAU SILO (Spéc: 5,3% - 6,3%)\n"
    if data['niveau_silo']:
        for item in data['niveau_silo']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"

    content += "\n4. HUMIDITÉ RÉSIDUELLE SÉCHOIR (Spéc: 0,1% - 1,5%)\n"
    if data['residuelle_sechoir']:
        for item in data['residuelle_sechoir']:
            time_str = item['time'].strftime('%H:%M') if item['time'] else 'N/A'
            status = "✓" if item['compliant'] else "✗"
            content += f"   {time_str} - {item['value']}% - {item['controller']} - {status}\n"
    else:
        content += "   Aucune mesure disponible\n"
    
    # Create a simple text file as fallback
    buffer = io.BytesIO(content.encode('utf-8'))
    filename = f"Fiche_Humidite_{data['date'].strftime('%Y%m%d')}.txt"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='text/plain'
    )